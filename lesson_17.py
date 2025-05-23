from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# wb = load_workbook('data_2.xlsx', read_only=True)
# ws = wb.active
# print(wb.sheetnames)
#
# sheet = wb['Лист1']
#
# value = sheet['B2'].value
#
# print(value)
#
# for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=5):
#     for cell in row:
#         print(cell.coordinate, cell.value)
#
#
#
# sheet['A3'] = "Новое значение"
# wb.save('data_2.xlsx')
#
# value = sheet['A3'].value
# print(value)


# wb = load_workbook('data.xlsx')
# ws = wb.active
# sheet = wb['Лист1']
# sheet.append(['Name', 'Age', 'City'])
# sheet.append(['Azat', '22', 'Bishkek'])
# sheet.append(['Dastan', '33', 'Karakol'])
# sheet.append(['Adilet', '22', 'Naryn'])
# wb.save('data.xlsx')


wb = Workbook()
ws = wb.active
ws.title = "Sheet"

ws['A1'] = "Name"
ws['B1'] = "Age"
ws['C1'] = "City"

data = [
    ['Azat', 22, 'Bishkek'],
    ['Dastan', 33, 'Karakol'],
    ['Adilet', 299, 'Naryn']
]

start_row = 2
for person in data:
    ws[f"A{start_row}"] = person[0]
    ws[f"B{start_row}"] = person[1]
    ws[f"C{start_row}"] = person[2]
    start_row += 1


ws['A1'].font = Font(name='Arial', size=14, bold=True, color='FF0000')
ws['A1'].alignment = Alignment(horizontal='center')
ws['A2'].fill = PatternFill("solid", fgColor="FF0000")

ws['B5'] = "=AVERAGE(B2:B3:B4)"
wb.save(filename="test.xlsx")






